import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import pandas as pd
from top_bar import TopBar
import customtkinter as ctk
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from costs import size_cost_assignments
from size_steel_assignments import steel_weight_assignments
import json
import os 
import datetime
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
import base64
import ctypes
from pathlib import Path

sheet_index = 0
window = None
image_var = None
image_frame = None
image_listbox = None
preview_label = None
reinforcement_combobox = None
file_created = None
images = []
cumulative_cost = 0

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(r"E:\Projects\Tendering\Figma\build\build\assets\frame0")

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)

def find_matching_size(selected_size):
    # Iterate through the keys in size_cost_assignments to find a match
    for size_key in size_cost_assignments.keys():
        if selected_size in size_key:
            return size_key
    return None

def retrieve_steel_info(selected_size, selected_image, selected_reinforcement_type, steel_weight_assignments):
    # Find the matching size based on user input
    matched_size = find_matching_size(selected_size)

    if matched_size is None:
        print(f"No matching size found for {selected_size}")
        return None

    # Retrieve size data based on matched_size
    size_data = steel_weight_assignments.get(matched_size, {})

    if not size_data:
        print(f"No data found for size {matched_size}")
        return None

    # Retrieve type data based on selected_image
    type_data = size_data.get(selected_image, {})

    if not type_data:
        print(f"No data found for type {selected_image} in size {matched_size}")
        return None

    # Retrieve steel weight based on selected_reinforcement_type
    steel_weight = type_data.get(selected_reinforcement_type, {}).get("weight(Kg)", 0)

    if steel_weight == 0:
        print(f"No steel weight found for reinforcement type {selected_reinforcement_type} in size {matched_size}, type {selected_image}")

    print(f"reinforcement, type, size: {selected_reinforcement_type, selected_image, matched_size}")

    # Debug statements for steel weight retrieval
    print(f"Selected Steel Weight (Kg): {steel_weight}")

    return steel_weight


def calculate_final_price(selected_size, length, quantity, dollar_price, steel_price, selected_image, selected_reinforcement_type, selected_material, updated_decking_value, updated_cladding_value, updated_pergola_value, shipping_price_value):
    # Check if the selected_size is in the size_cost_assignments dictionary
    if selected_size not in size_cost_assignments:
        print(f"Size '{selected_size}' not found in size_cost_assignments. Cannot calculate steel price.")
        return None, None, None  # or you can return some default values or handle it as needed

    # Assuming cost_value is the initial cost_value
    cost_value = size_cost_assignments[selected_size].get("price", 0)

    # Calculate weight_value by multiplying size_cost_assignments['weight'] with length and quantity
    weight_value = size_cost_assignments[selected_size].get("weight", 0) * length * quantity #136.62

    weight_wpc = size_cost_assignments[selected_size].get("weight", 0)

    # Calculate the cost by multiplying cost_value with length
    cost = (cost_value + ((weight_wpc / 22000) * shipping_price_value)) * quantity * length
    print(f"Cost: {round(cost, 1)}")  # Round to 2 decimal places
    print(f"Length: {length}")
    print(f"Weight: {weight_wpc}")
    print(f"shipping: {shipping_price_value}")

    # Calculate the total before profit in EGP
    total_before_profit = (cost * 1.7 * 1.14) + (cost * 0.1)
    print(f"Total before profit (USD): {round(total_before_profit, 0)}")  # Round to 2 decimal places\
    print(f"shipping: {total_before_profit}")

    # Assuming dollar_price is the USD to EGP conversion rate
    price_after_conversion = total_before_profit * dollar_price
    print(f"Total after conversion (EGP): {round(price_after_conversion, 0)}")  # Round to 2 decimal places

    # Calculate the final profit price in EGP
    wpc_price = price_after_conversion / 0.7
    print(f"WPC Selling Price (EGP): {round(wpc_price, 0)}")  # Round to 2 decimal places

    # Call retrieve_steel_info with the required arguments
    steel_weight = retrieve_steel_info(selected_size, selected_image, selected_reinforcement_type, steel_weight_assignments)

    # Call retrieve_steel_info with the required arguments only if selected_material is not Decking or Cladding
    if selected_material not in ["Decking", "Cladding"]:
        steel_weight = retrieve_steel_info(selected_size, selected_image, selected_reinforcement_type, steel_weight_assignments)

        # Check if steel_weight is None (indicating size not found in steel_weight_assignments)
        if steel_weight is None:
            print(f"Size '{selected_size}' not found in steel_weight_assignments. Cannot calculate steel price.")
            return wpc_price, None, weight_value, install_price  # or handle it as needed

    # If selected_material is Decking or Cladding, set steel_weight to 0
    else:
        steel_weight = 0

    # Calculate Steel Selling Price
    steel_cost = (length * steel_weight * steel_price * quantity) / 0.7
    print(f"Steel Weight (Kg): {steel_weight}")
    print(f"Steel Selling Price (EGP): {round(steel_cost, 0)}")  # Round to 2 decimal places

    # Calculate install_price based on the selected material
    if selected_material == "Pergola":
        installment_factor = updated_pergola_value
        print(f"installment factor: {installment_factor}")
    elif selected_material == "Decking":
        installment_factor = updated_decking_value
        print(f"installment factor: {installment_factor}")
    else:
        installment_factor = updated_cladding_value
        print(f"installment factor: {installment_factor}")

    if selected_material in ["Decking", "Cladding"]:
        install_price = round((length * installment_factor) / 0.7, 1)
        print(f"installment lenghth price: ({weight_value} * {installment_factor}) / 0.7 = {install_price}")
    else:
        install_price = round((weight_value * installment_factor) / 0.7, 1)
        print(f"installment weight price: ({weight_value} * {installment_factor}) / 0.7 = {install_price}")

    print(f"Values before returning: wpc_price={wpc_price}, steel_cost={steel_cost}, weight_value={weight_value}, install_price={install_price}")
    return wpc_price, steel_cost, weight_value, install_price, cost


def decrypt_config_file(file_path, encoded_key):
    with open(file_path, 'rb') as f:
        encrypted_data = f.read()

    try:
        # Decode the base64-encoded key
        key = base64.b64decode(encoded_key)

        # Ensure the key length is correct (16 bytes for AES-128)
        if len(key) != 16:
            raise ValueError("Invalid key length")

        cipher = Cipher(algorithms.AES(key), modes.ECB(), backend=default_backend())
        decryptor = cipher.decryptor()
        decrypted_data = decryptor.update(encrypted_data) + decryptor.finalize()

        # Remove PKCS7 padding
        padding_length = decrypted_data[-1]
        decrypted_data = decrypted_data[:-padding_length]

        # Parse the JSON data
        config_data = json.loads(decrypted_data.decode('utf-8'))

        # Debug print statement to output the entire decrypted configuration data
        print("Decrypted Configuration Data:", config_data)

        return config_data

    except Exception as e:
        print(f"Error decrypting: {e}")
        return None

def update_window_title():
    global window, excel_project_name, last_created_sheet
    window.title(f"W&M Pergola calculator (File Name: {excel_project_name})")

def resize_image(image_path, target_width, target_height):
    if not image_path:
        return None  # Return None for an empty image path

    original_image = Image.open(image_path)
    original_width, original_height = original_image.size

    # Calculate the aspect ratio
    aspect_ratio = original_width / original_height

    # Calculate the target size while maintaining the aspect ratio
    if aspect_ratio > 1:
        # Landscape orientation
        new_width = target_width
        new_height = int(target_width / aspect_ratio)
    else:
        # Portrait or square orientation
        new_width = int(target_height * aspect_ratio)
        new_height = target_height

    # Resize the image
    resized_image = original_image.resize((new_width, new_height))

    # Convert the resized PIL Image to a PhotoImage
    tk_image = ImageTk.PhotoImage(resized_image)

    return tk_image

def update_options(event):
    global image_frame, image_var, image_listbox, images, preview_label, sheet_indicator_label, reinforcement_combobox, size_image_dict


    selected_material = material_var.get()

    # Update size options based on the selected material
    material_size_dict = {
        "Pergola": ["","15 x 15", "12 x 12", "20 x 20", "10 x 10", "8 x 8", "5 x 5", "7 x 7", "20 x 10", "15 x 5", "20 x 15", "20 x 8", "5 x 10", "16 x 8", "9 x 7", "5 x 3"],
        "Decking": ["","14.6 x 2.2", "14 x 2.5", "10 x 2.5", "14 x 2.2", "14 x 2.5", "14.6 x 2.3", "27.9 x 1.9"],
        "Cladding": ["","15 x 1.3", "15.6 x 2.1", "14 x 1.2", "17.7 x 2.8", "7 x 1.6", "9.8 x 1.3", "21.9 x 2.6"]
    }
    sizes = material_size_dict.get(selected_material, [])

    selected_size = size_var.get()  # Get the currently selected size

    # Clear the existing size options
    size_menu['values'] = sizes

    if selected_size not in sizes:
        size_var.set(sizes[0] if sizes else "")  # Set default value if the current size is not in the new list

    # Update image options based on the selected size
    size_image_dict = {
        #Pergola
        "15 x 15": ["images/Picture3.png", "images/Picture9.png", "images/Picture10.png"],
        "12 x 12": ["images/Picture4.png"],
        "20 x 20": ["images/Picture1.png"],
        "10 x 10": ["images/Picture2.png", "images/Picture8.png"],
        "8 x 8": ["images/Picture5.png"],
        "5 x 5": ["images/Picture6.png"],
        "7 x 7": ["images/Picture7.png"],
        "20 x 10": ["images/Picture11.png"],
        "15 x 5": ["images/Picture12.png"],
        "20 x 15": ["images/Picture13.png"],
        "20 x 8": ["images/Picture14.png", "images/Picture18.png"],
        "5 x 10": ["images/Picture15.png", "images/Picture19.png"],
        "16 x 8": ["images/Picture16.png"],
        "9 x 7": ["images/Picture17.png"],
        "5 x 3": ["images/Picture35.png", "images/Picture36.png"],
        #Decking
        "14.6 x 2.2": ["images/Picture23.png"],
        "14 x 2.5": ["images/Picture24.png"],
        "10 x 2.5": ["images/Picture25.png"],
        "14 x 2.2": ["images/Picture26.png", "images/Picture28.png"],
        "14.6 x 2.3": ["images/Picture32.png", "images/Picture33.png"],
        "27.9 x 1.9": ["images/Picture34.png"],
        "15 x 2.5": ["images/Picture38.png"],
        "14.3 x 2.25": ["images/Picture39.png"],
        "14 x 2": ["images/Picture40.png"],
        #Cladding
        "15 x 1.3": ["images/Picture20.png"],
        "15.6 x 2.1": ["images/Picture21.png"],
        "14 x 1.2": ["images/Picture22.png"],
        "17.7 x 2.8": ["images/Picture29.png"],
        "7 x 1.6": ["images/Picture30.png"],
        "9.8 x 1.3": ["images/Picture31.png"],
        "21.9 x 2.6": ["images/Picture37.png"],
    }
    images = size_image_dict.get(selected_size, [])

    # Clear the existing image options
    if image_frame:
        for widget in image_frame.winfo_children():
            widget.destroy()

    # Check if there are images for the selected size
    if images:
        # Create a listbox for selecting an image
        image_listbox = tk.Listbox(image_frame, selectmode=tk.SINGLE, bg='#D4FADD', fg='black', width=10, height=0, highlightcolor="#242424", font=("Verdana", 16))
        image_listbox.pack(side=tk.LEFT, fill=tk.BOTH)

        # Create TkImage objects for each image associated with the selected size
        photo_images = [resize_image(image_path, 250, 250) for image_path in images]

        # Add images to the listbox
        for i, ctimage in enumerate(photo_images):
            label = ttk.Label(image_listbox, image=ctimage, text=f"Type {i + 1}", compound=tk.LEFT)
            label.photo = ctimage  # Keep a reference to avoid garbage collection
            image_listbox.insert(tk.END, f"Type {i + 1}")

        # Create a preview label
        preview_label = ttk.Label(image_frame, text="Preview:")
        preview_label.pack(pady=5)

        # Update the preview label with the first image
        if photo_images:
            preview_image = resize_image(images[0], 250, 250)
            preview_label.configure(image=preview_image, text="")
            preview_label.image = preview_image  # Keep a reference to avoid garbage collection
        # Update the unit label for the length input based on the selected material
        length_unit_label.config(text="Length (meters):" if selected_material == "Pergola" else "Area (m²):")
    else:
        # No images available for the selected size
        image_listbox = tk.Listbox(image_frame, bg='#D4FADD', fg='black', width=10, height=0, highlightcolor="#1F6AA5", font=("Verdana", 16))
        image_listbox.pack(side=tk.LEFT, fill=tk.BOTH)
        image_listbox.insert(tk.END, "None")  # Set default value to "None"()

        # Create a preview label with a default image
        preview_label = ttk.Label(image_frame, text="Preview:")
        preview_label.pack(pady=5)

        default_preview_image = resize_image("images/icons/NA.jpg", 250, 250)  # Replace with the path to your default image
        preview_label.configure(image=default_preview_image, text="")
        preview_label.image = default_preview_image  # Keep a reference to avoid garbage collection

        # Make the image_listbox non-interactable
        image_listbox.config(state=tk.DISABLED)

        # Update the unit label for the length input based on the selected material
        length_unit_label.config(text="Length (meters):" if selected_material == "Pergola" else "Area (m²)")

    # Update the sheet indicator
    sheet_indicator_label.config(text="Sheet: " + last_created_sheet)

    # Update reinforcement options
    update_reinforcement_options()

    # Update component options based on the selected material
    components = [] if selected_material != "Pergola" else ["Column", "Main Beam", "Secondary Beam"]
    component_menu['values'] = components
    component_var.set(components[0] if components else "None")  # Set default value if available
    component_menu.config(state=tk.NORMAL if components else tk.DISABLED)  # Make it non-interactable if no components

    # Bind the listbox selection event if images are present
    if images:
        image_listbox.bind("<<ListboxSelect>>", update_preview)
        image_listbox.focus_set()  # Set focus to the listbox after selection

def update_reinforcement_options(*args):
    global reinforcement_combobox, reinforcement_var

    selected_option = reinforcement_var.get()

    if selected_option == "Yes":
        # Populate the combobox with options
        reinforcement_combobox['values'] = ["2mm U shape","3mm U shape", "2mm Box shape", "3mm Box shape"]

        # Set the combobox value to the first option
        if reinforcement_combobox['values']:
            reinforcement_combobox.set(reinforcement_combobox['values'][0])

        # Make the combobox interactable
        reinforcement_combobox.config(state=tk.NORMAL)
    else:
        # Clear the combobox values for "No" option
        reinforcement_combobox.set("None")  # Clear the combobox selection
        reinforcement_combobox['values'] = []  # Clear the values

        # Make the combobox non-interactable
        reinforcement_combobox.config(state=tk.DISABLED)

def update_preview(event):
    global image_listbox, images, preview_label

    selected_index = image_listbox.curselection()[0] if image_listbox.curselection() else -1
    selected_image = images[selected_index] if selected_index != -1 else ""

    # Update the preview label
    if selected_index != -1:
        preview_image = resize_image(selected_image, 250, 250)
        preview_label.configure(image=preview_image, text="")
        preview_label.image = preview_image  # Keep a reference to avoid garbage collection

        # Set focus to the image_listbox
        image_listbox.focus_set()


def save_to_excel(data, excel_project_name, sheet_name, config_data, cost, cumulative_cost, is_retry=False):
    import pandas as pd
    print("Starting save_to_excel function.")

    try:
        # Read existing data from the Excel file if it exists
        existing_data = pd.read_excel(excel_project_name, sheet_name=None)
        print(f"Existing data read from {excel_project_name}.")
    except FileNotFoundError:
        print(f"File {excel_project_name} not found. Creating a new file with headers.")
        # Create a new file with just the headers (columns from data)
        pd.DataFrame(columns=data.columns).to_excel(excel_project_name, sheet_name=sheet_name, index=False)
        if not is_retry:
            # Recall the function to ensure headers are created and file is established
            print("Recalling save_to_excel after file creation.")
            save_to_excel(data, excel_project_name, sheet_name, config_data, cost, cumulative_cost, is_retry=True)
            return

    if is_retry:
        # If retrying, reload the data to get the updated existing_data including only the headers
        existing_data = pd.read_excel(excel_project_name, sheet_name=None)

    if not is_retry:
        # Only combine data on the initial call, not on retry
        combined_data = pd.concat([existing_data.get(sheet_name, pd.DataFrame()), data], ignore_index=True)
        print(f"Data combined. New data rows: {len(data)}. Combined data rows: {len(combined_data)}.")
    else:
        # On retry, the existing_data should already have the correct structure but be empty of entries
        combined_data = data

    # Check if the total row is already present in the combined data
    total_row_index = combined_data[combined_data['Material'] == 'Total'].index
    if not total_row_index.empty:
        # Remove the existing total row(s)
        combined_data = combined_data.drop(total_row_index)

    # Calculate totals and append the total row
    total_weight = round(combined_data["WPC Weight"].sum(), 1)
    total_cost = round(combined_data["WPC Selling Price"].sum(), 1)
    total_steel = round(combined_data["Steel Selling Price"].sum(), 1)
    total_install = round(combined_data["Install Price"].sum(), 1)

    total_row = pd.DataFrame({
        "Material": ["Total"],
        "Component": [""],
        "Size": [""],
        "Image": [""],
        "Quantity": [""],
        "Length": [""],  
        "Reinforcement": [""],
        "Reinforcement Type": [""],
        "WPC Weight": [f"{total_weight:,.1f} Kg"],
        "WPC Selling Price": [f"{total_cost:,.1f} EGP"],
        "Install Price": [f"{total_install:,.1f} EGP"],
        "Steel Selling Price": [f"{total_steel:,.1f} EGP"],
        "Final": [f"{round(total_install + total_cost + total_steel, 1):,.1f} EGP"],
    })

    # Append the total row
    combined_data = pd.concat([combined_data, total_row], ignore_index=True)

    # Check if the "last updated" rows already exist and remove them
    last_updated_row_indices = combined_data[combined_data['Material'].isin(['USD Price', 'Steel Price', 'Updated Pergola', 'Updated Decking', 'Updated Cladding', 'Shipping Fees', 'Last updated', 'Date', '-'])].index
    if not last_updated_row_indices.empty:
        combined_data = combined_data.drop(last_updated_row_indices)

    additional_rows = [
        {"Material": "Last updated", "Component": "Date"},
        {"Material": "USD Price", "Component": config_data['usd_price']['last_updated']},
        {"Material": "Steel Price", "Component": config_data['steel_price']['last_updated']},
        {"Material": "Updated Pergola", "Component": config_data['updated_pergola']['last_updated']},
        {"Material": "Updated Decking", "Component": config_data['updated_decking']['last_updated']},
        {"Material": "Updated Cladding", "Component": config_data['updated_cladding']['last_updated']},
        {"Material": "Shipping Fees", "Component": config_data['shipping']['last_updated']},
        {"Material": "-", "Component": cumulative_cost},
    ]

    # Convert the additional rows dictionary to a DataFrame
    additional_rows_df = pd.DataFrame(additional_rows)

    # Append the additional rows DataFrame to the combined data DataFrame
    combined_data = pd.concat([combined_data, additional_rows_df], ignore_index=True)

    # Write the combined data to the Excel file
    with pd.ExcelWriter(excel_project_name, engine='xlsxwriter') as writer:
        for existing_sheet_name, existing_sheet_data in existing_data.items():
            if existing_sheet_name == sheet_name:
                combined_data.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                existing_sheet_data.to_excel(writer, sheet_name=existing_sheet_name, index=False)

    print("save_to_excel function completed.")
            


    

# Initialize last_created_sheet to a default value
last_created_sheet = "Pergola_1"


def create_new_sheet():
    global last_created_sheet, excel_project_name, sheet_indicator_label

    try:
        # Read existing data from the Excel file if it exists
        existing_data = pd.read_excel(excel_project_name, sheet_name=None)

        # Generate a new sheet name automatically (e.g., Pergola_1, Pergola_2, ...)
        new_sheet_index = len(existing_data) + 1
        new_sheet_name = f"Pergola_{new_sheet_index}"

        # Create a new sheet with an empty DataFrame
        new_data = pd.DataFrame()

        # Save the new sheet to the Excel file
        with pd.ExcelWriter(excel_project_name, engine='xlsxwriter') as writer:
            # Write the new sheet to the Excel file
            new_data.to_excel(writer, sheet_name=new_sheet_name, index=False)

            # Write the existing sheets back to the Excel file
            for sheet_name, sheet_data in existing_data.items():
                sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)

        # Store the last created sheet name
        last_created_sheet = new_sheet_name

        # Update the sheet indicator
        sheet_indicator_label.config(text="Sheet: " + last_created_sheet)

        # Show a message indicating that a new sheet has been created
        messagebox.showinfo("New Sheet Created", "A new sheet has been created successfully!")

    except FileNotFoundError:
        # If the file doesn't exist, create a new one with no default data
        new_data = pd.DataFrame()
        new_data.to_excel(excel_project_name, index=False, sheet_name="Pergola_1")

        # Store the last created sheet name
        last_created_sheet = "Pergola_1"

        # Update the sheet indicator
        sheet_indicator_label.config(text="Sheet: " + last_created_sheet)

        # Show a message indicating that a new sheet has been created
        messagebox.showinfo("New Sheet Created", "A new sheet has been created successfully!")


# Example function to check if excel_project_name is defined
def check_excel_defined():
    if excel_project_name is None:
        messagebox.showerror("Error", "No Excel file loaded. Please load a file first.")
        return False
    return True

# Function to move to the next sheet
def move_to_next_sheet():
    global last_created_sheet, sheet_indicator_label, sheet_index, excel_project_name

    # Check if excel_project_name is defined
    if not check_excel_defined():
        return

    # Implement logic to move to the next sheet
    sheet_names = get_sheet_names()

    # Show an error message if there are no sheets available
    if not sheet_names:
        messagebox.showerror("Error", "No sheets available in the Excel file.")
        return

    sheet_index = (sheet_index + 1) % len(sheet_names)
    last_created_sheet = sheet_names[sheet_index]

    # Update the sheet indicator
    sheet_indicator_label.config(text="Sheet: " + last_created_sheet)

# Function to move to the previous sheet
def move_to_previous_sheet():
    global last_created_sheet, sheet_indicator_label, sheet_index, excel_project_name

    # Check if excel_project_name is defined
    if not check_excel_defined():
        return

    # Implement logic to move to the previous sheet
    sheet_names = get_sheet_names()

    # Show an error message if there are no sheets available
    if not sheet_names:
        messagebox.showerror("Error", "No sheets available in the Excel file.")
        return

    sheet_index = (sheet_index - 1) % len(sheet_names)
    last_created_sheet = sheet_names[sheet_index]

    # Update the sheet indicator
    sheet_indicator_label.config(text="Sheet: " + last_created_sheet)

# Function to get sheet names
def get_sheet_names():
    # Implement logic to get a list of sheet names from the Excel file
    try:
        existing_data = pd.read_excel(excel_project_name, sheet_name=None)
        return list(existing_data.keys())
    except FileNotFoundError:
        return []
    
# Example usage:
config_file_path = "config_encrypted.json"  # Replace with your actual file path

# Encode the actual key
actual_key = b'StaticKey16Bytes'
encoded_key = base64.b64encode(actual_key).decode('utf-8')

# Decrypt the configuration file
config_data = decrypt_config_file(config_file_path, encoded_key)

# Retrieve and print the dollar price
dollar_price_value = config_data.get("usd_price", 1.0)  # Default value is 1.0 if not found in the config

# Retrieve and print the steel price
steel_price_value = config_data.get("steel_price", 1.0)  # Default value is 1.0 if not found in the config

# Retrieve and print the updated pergola value
updated_pergola_value = config_data.get("updated_decking", 1)

# Retrieve and print the updated cladding value
updated_cladding_value = config_data.get("updated_pergola", 1)

# Retrieve and print the updated decking value
updated_decking_value = config_data.get("updated_cladding", 1)



def on_button_click():
    global image_listbox, images, preview_label, excel_project_name, last_created_sheet, size_image_dict, file_created, total, cumulative_cost

    # Check if the file has already been created
    if not file_created:
        # Open a file dialog to get the Excel project name and location
        excel_project_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        # Check if the user clicked cancel or provided an empty project name
        if not excel_project_name:
            return
        
        # Get the current date
        current_date = datetime.datetime.now().strftime("(%d-%m-%Y)")
        
        # Split the path and the file name
        path, filename = os.path.split(excel_project_name)
        # Remove the extension from the filename
        basename, ext = os.path.splitext(filename)
        
        # Append the date to the filename
        modified_basename = f"{basename} {current_date}"
        
        # Combine everything back to get the final path with the modified name
        excel_project_name = os.path.join(path, f"{modified_basename}{ext}")

        # Set the boolean variable to True after creating the file
        file_created = True

    selected_material = material_var.get()
    selected_component = component_var.get()
    selected_size = size_var.get()
    quantity = quantity_var.get()
    length = length_var.get()

    # Input validation
    if not length or not quantity or not selected_size:
        label.config(text="Please fill in all required fields.", font=("Verdana", 12), foreground="red")
        return

    try:
        # Check if length and quantity are valid numeric values
        length = float(length)
        quantity = int(quantity)
    except ValueError:
        label.config(text="Please enter valid numeric values for Length and Quantity.", font=("Verdana", 12), foreground="red")
        return

    if length <= 0 or quantity <= 0:
        label.config(text="Please enter positive values for Length and Quantity.", font=("Verdana", 12), foreground="red")
        return

    selected_index = image_listbox.curselection()[0] if image_listbox.curselection() else -1
    selected_image = f"Type {selected_index + 1}" if selected_index != -1 else ""

    # Input validation: Check if the selected size has images and if an image is selected
    if selected_size in size_image_dict and not selected_image:
        label.config(text="Please select an image for the selected size.", font=("Verdana", 12), foreground="red")
        return
    
    # Get the state of the reinforcement radio buttons
    reinforcement_state = reinforcement_var.get()
    # Get the selected reinforcement type from the combobox
    selected_reinforcement_type = reinforcement_combobox_var.get()

    # Input validation for reinforcement only if the reinforcement radio box is "Yes"
    if reinforcement_state == "Yes" and not selected_reinforcement_type:
        label.config(text="Please select a reinforcement material.", font=("Verdana", 12), foreground="red")
        return

        # Extract the values from the config_data before passing them to the function
    dollar_price_value = config_data['usd_price']['value']
    steel_price_value = config_data['steel_price']['value']
    updated_decking_value = config_data['updated_decking']['value']
    updated_cladding_value = config_data['updated_cladding']['value']
    updated_pergola_value = config_data['updated_pergola']['value']
    shipping_price_value = config_data['shipping']['value']

    # Now, call the function with these extracted values
    wpc_price, steel_cost, weight_value, install_price, cost = calculate_final_price(
        selected_size, length, quantity,
        dollar_price_value, steel_price_value,
        selected_image, selected_reinforcement_type, selected_material,
        updated_decking_value, updated_cladding_value, updated_pergola_value, shipping_price_value
    )
    cumulative_cost += cost

    # Determine the unit based on the selected material
    if selected_material == "Pergola":
        length_unit = "M"
    else:
        # Assuming 'Decking' and 'Cladding' are the other options
        length_unit = "M^2"

    # Create a DataFrame with the current data, including the reinforcement information and cost
    data = pd.DataFrame({
        "Material": [selected_material],
        "Component": [selected_component],
        "Size": [selected_size],
        "Color": [color_var.get()],  # Add the selected color
        "Grain": [grain_var.get()],  # Add the grain selection
        "WPC Weight": [weight_value],
        "Image": [selected_image],  # Store the selected index (+1 to match your "Type" numbering)
        "Quantity": [quantity],
        "Length": [f"{length} {length_unit}"],  # Append the appropriate unit to the length value
        "Reinforcement": [reinforcement_state],  # Add the reinforcement information to the DataFrame
        "Reinforcement Type": [selected_reinforcement_type],
        "WPC Selling Price": [wpc_price],
        "Steel Selling Price": [steel_cost],
        "Install Price": [install_price]
    })
    
    # Modify the result_text construction
    result_text = f"Successfully generated component: {selected_component}, Size: {selected_size}"

    label.config(text=result_text)

    save_to_excel(data, excel_project_name, last_created_sheet, config_data, cost, cumulative_cost)
    update_window_title()

        # Open the existing workbook
    workbook = load_workbook(excel_project_name)
    worksheet = workbook.active
        # Find the last row number
    last_row_number = worksheet.max_row

    # Apply white font color to the entire last row
    for col in range(1, worksheet.max_column + 1):  # Iterating through each column
        cell = worksheet.cell(row=last_row_number, column=col)
        cell.font = Font(color="FFFFFF")  # Setting font color to white

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        # Access the sheet
        sheet = workbook[sheet_name]

        # Iterate through columns and adjust the column width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name (e.g., 'A', 'B', 'C', ...)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

                if get_column_letter(cell.column) in ["L", "M", "N"] and cell.row != 1:
                    # Check if the cell's row contains the "Total" label in the first column
                    if sheet[f"A{cell.row}"].value != "Total":
                        cell.font = Font(color="FFFFFF")  # Assuming you want black text, use "000000" for black

            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Save the changes to the Excel file
    workbook.save(excel_project_name)
    workbook.close()
        # Play a sound alert
    play_sound_alert()

def play_sound_alert():
    # Use ctypes to play the default Windows alert sound
    ctypes.windll.user32.MessageBeep(0)


# Check if the user clicked cancel or provided an empty project name
    
# Initialize the main application window
window = ctk.CTk()
window.iconbitmap("images/icons/logo.ico")  # Replace with the path to your .ico file
window.title("W&M Tendering")
window.geometry("1280x720")
window.configure(bg="#242424")  # Assuming ctk.CTk has a configure method for background color
window._set_appearance_mode("Dark")  # Custom TKinter specific for dark mode

# Create and apply a theme with ttk for additional widgets that ctk may not cover
style = ttk.Style()
style.theme_use('alt')  # Assuming 'alt' is a valid theme

# Create the top bar
top_bar = TopBar(window)

# Define general style configurations
accent_color = "#014421"
background_color = "#242424"
text_color = "white"
font = ('Arial', 12)

# Main content frame - this will contain most of your widgets
main_frame = ttk.Frame(window, style="ImageFrame.TFrame")
main_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
window.columnconfigure(0, weight=1)
window.rowconfigure(0, weight=1)

# Example: Material Options
materials = ["Pergola", "Decking", "Cladding"]
material_label = ttk.Label(main_frame, text="Item:", background=background_color, foreground=text_color, font=("Verdana", 12))
material_label.grid(row=0, column=0, padx=5, pady=5)
material_var = tk.StringVar(main_frame)
material_var.set(materials[0])  # Default selection
material_menu = ttk.Combobox(main_frame, textvariable=material_var, state="readonly", values=materials, font=("Verdana", 12), width=10)
material_menu.grid(row=0, column=1, padx=5, pady=5)

# Component Options
components = ["Column", "Main Beam", "Secondary Beam"]
component_label = ttk.Label(main_frame, text="Component:", background="#242424", foreground="white", font=("Verdana", 12))
component_label.grid(row=0, column=4, padx=5)
component_var = tk.StringVar(main_frame)
component_var.set(components[0])  # Set default value
component_menu = ttk.Combobox(main_frame, textvariable=component_var, state="readonly", values=components, font=("Verdana", 12), width=10)
component_menu.grid(row=0, column=5, padx=5)

# Length Input
length_label = ttk.Label(main_frame, text="Length/Area:", background="#242424", foreground="white", font=("Verdana", 12))
length_label.grid(row=1, column=2, padx=5)

length_var = tk.DoubleVar(main_frame)  # Use DoubleVar for decimal values
length_entry = ttk.Entry(main_frame, textvariable=length_var, font=("Verdana", 12), width=12)
length_entry.grid(row=1, column=3, padx=5)

# Create a label for unit
length_unit_label = ttk.Label(main_frame, text="Length (meters)", background="#242424", foreground="white", font=("Verdana", 12))
length_unit_label.grid(row=1, column=4, padx=5)

# Size Options
size_label = ttk.Label(main_frame, text="Size:", background="#242424", foreground="white", font=("Verdana", 12))
size_label.grid(row=0, column=2, padx=10, pady=40)
size_var = tk.StringVar(main_frame)
size_menu = ttk.Combobox(main_frame, textvariable=size_var, state="readonly", values=[], font=("Verdana", 12), width=10)
size_menu.grid(row=0, column=3, padx=10, pady=40)

# Create a custom style for the frame
style = ttk.Style()
style.configure("ImageFrame.TFrame", background="#242424")  # Set the background color

# Image Options
image_frame = ttk.Frame(window, style="ImageFrame.TFrame")
image_frame.place(x=200, y=350)  # Adjust x and y values as needed
image_var = tk.StringVar(window)
image_var.set("")  # Set default value


# Quantity Counter
quantity_label = ttk.Label(main_frame, text="Quantity:", background="#242424", foreground="white", font=("Verdana", 12))
quantity_label.grid(row=1, column=0, padx=10, pady=40)
quantity_var = tk.StringVar(main_frame)
quantity_var.set(1)  # Set default value
quantity_spinbox = ttk.Spinbox(main_frame, from_=1, to=900, textvariable=quantity_var, font=("Verdana", 12), width=10)
quantity_spinbox.grid(row=1, column=1, padx=10, pady=40)

# Reinforcement Radio Buttons
reinforcement_label = ttk.Label(main_frame, text="Reinforcement:", background="#242424", foreground="white", font=("Verdana", 12))
reinforcement_label.grid(row=0, column=6, pady=40, padx=20)

reinforcement_var = tk.StringVar(main_frame, value="Yes")  # Set the default value to "Yes"

# Bind the event to update options
reinforcement_var.trace_add("write", update_reinforcement_options)

# Create a frame for the radio buttons
reinforcement_frame = ttk.Frame(main_frame, style="ImageFrame.TFrame")
reinforcement_frame.grid(row=0, column=7, pady=5, padx=20)

reinforcement_yes_radio = tk.Radiobutton(reinforcement_frame, text="Yes", variable=reinforcement_var, value="Yes", background="#242424", foreground="white", selectcolor="#242424", font=("Verdana", 12))
reinforcement_yes_radio.grid(row=0, column=0, padx=5)

reinforcement_no_radio = tk.Radiobutton(reinforcement_frame, text="No", variable=reinforcement_var, value="No", background="#242424", foreground="white", selectcolor="#242424", font=("Verdana", 12))
reinforcement_no_radio.grid(row=0, column=1, padx=5)

# Bind the event to update options
reinforcement_frame.bind("<Configure>", update_reinforcement_options)

# Reinforcement Combobox
reinforcement_combobox_label = tk.Label(main_frame, text="Type:", background="#242424", foreground="white", font=("Verdana", 12))
reinforcement_combobox_label.grid(row=1, column=5, pady=5, padx=20)

reinforcement_combobox_var = tk.StringVar(main_frame)
reinforcement_combobox = ttk.Combobox(main_frame, textvariable=reinforcement_combobox_var, state="readonly", values=[], font=("Verdana", 12), width=13)
reinforcement_combobox.grid(row=1, column=6, pady=20)

# Color Options
color_options = ["P801", "P8", "P7", "P6", "P501", "P5", "P4", "P301", "P3", "P201", "P1"]
color_label = ttk.Label(main_frame, text="Color:", background="#242424", foreground="white", font=("Verdana", 12))
color_label.grid(row=2, column=0, padx=10, pady=20)
color_var = tk.StringVar(main_frame)
color_combobox = ttk.Combobox(main_frame, textvariable=color_var, state="readonly", values=color_options, font=("Verdana", 12), width=10)
color_combobox.grid(row=2, column=1, padx=10, pady=20)
color_combobox.set(color_options[0])  # Set default value

# Grain Selection Label
grain_label = ttk.Label(main_frame, text="Grain:", background="#242424", foreground="white", font=("Verdana", 12))
grain_label.grid(row=2, column=2, padx=10, pady=20)
grain_var = tk.StringVar(value="No")  # Default to "No"
grain_yes_radio = tk.Radiobutton(main_frame, text="Yes", variable=grain_var, value="Yes", background="#242424", foreground="white", selectcolor="#242424", font=("Verdana", 12))
grain_yes_radio.grid(row=2, column=3, padx=5, pady=20)
grain_no_radio = tk.Radiobutton(main_frame, text="No", variable=grain_var, value="No", background="#242424", foreground="white", selectcolor="#242424", font=("Verdana", 12))
grain_no_radio.grid(row=2, column=4, padx=5, pady=20)
    
# Bind the event to update options
material_menu.bind("<<ComboboxSelected>>", update_options)
size_menu.bind("<<ComboboxSelected>>", update_options)

# Result label
label = ttk.Label(main_frame, text="", background="#242424", foreground="white")
label.grid(row=10, column=5, pady=10)


# Bottom bar frame setup
bottom_bar_frame = ttk.Frame(window, style="ImageFrame.TFrame")
bottom_bar_frame.grid(row=1, column=0, sticky="ew", columnspan=3)
window.rowconfigure(1, weight=0)

style = ttk.Style()
style.configure("Nav.TFrame", background="#242424")

# Modify button creation by placing them in a new frame in the bottom_bar_frame for better alignment
buttons_frame = ttk.Frame(bottom_bar_frame, style="Nav.TFrame")  # Create a frame for buttons
buttons_frame.pack(side="right", padx=10)  # Pack it on the right side of the bottom_bar_frame

# Adjust "Calculate" button to be in buttons_frame
calculate_button = ctk.CTkButton(buttons_frame, text="Calculate", command=on_button_click, fg_color="#014421", hover_color="#4D7C63", font=("Verdana", 20), bg_color="#242424")
calculate_button.pack(side="top", pady=(0, 5), padx=20)  # Stack the button above, with padding below

# Adjust "End / New Item" button to be in buttons_frame
end_new_item_button = ctk.CTkButton(buttons_frame, text="End / New Item", command=create_new_sheet, fg_color="#014421", hover_color="#4D7C63", font=("Verdana", 20), bg_color="#242424")
end_new_item_button.pack(side="top", padx=20)  # Stack it above or below the Calculate button, as required


# Sheet indicator in the bottom bar
sheet_indicator_label = ttk.Label(bottom_bar_frame, text="Sheet: last_created_sheet", background="#242424", foreground="white", font=("Verdana", 12))
sheet_indicator_label.pack(side="left", padx=10)


# Navigation buttons container in the bottom bar
nav_buttons_frame = ttk.Frame(bottom_bar_frame, style="Nav.TFrame")  # Using a style if needed or just omit the style argument
nav_buttons_frame.pack(side="left", padx=10)  # Removed expand=True, fill='x' and used side="left"

# Create left arrow button within nav_buttons_frame
left_arrow_button = ctk.CTkButton(nav_buttons_frame, text="◀", command=move_to_previous_sheet, fg_color="#014421", bg_color="#242424")
left_arrow_button.pack(side="left", padx=5)  # Removed expand=True

# Create right arrow button within nav_buttons_frame
right_arrow_button = ctk.CTkButton(nav_buttons_frame, text="▶", command=move_to_next_sheet, fg_color="#014421", bg_color="#242424")
right_arrow_button.pack(side="left", padx=5)  # Changed side to "left" and removed expand=True

# Set the font size for navigation buttons
style = ttk.Style()
style.configure("Nav.TButton", font=('Verdana', 12), background="#242424")  # Adjust the font size as needed

# Start the Tkinter event loop
window.mainloop()
